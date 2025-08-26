import pandas as pd
import numpy as np
from openpyxl import Workbook
import datetime


def trapezoidal_mf(x, a, b, c, d):
    """梯形隶属函数：a=左底点，b=左顶点，c=右顶点，d=右底点"""
    if x <= a:
        return 0.0
    elif a < x <= b:
        return (x - a) / (b - a)
    elif b < x <= c:
        return 1.0
    elif c < x <= d:
        return (d - x) / (d - c)
    else:
        return 0.0


def entropy_weight(data):
    """熵权法计算客观权重：输入n行（样本）×m列（指标）的数据矩阵，输出m个指标的权重"""
    data = np.array(data)
    n, m = data.shape  # n=样本数，m=指标数
    if n < 2:
        raise ValueError("计算权重至少需要2个样本（行）数据")

    # 数据标准化（正向化）
    normalized_data = np.zeros_like(data)
    for j in range(m):
        col = data[:, j]
        min_val = np.min(col)
        max_val = np.max(col)
        if max_val != min_val:
            normalized_data[:, j] = (col - min_val) / (max_val - min_val)
        else:
            normalized_data[:, j] = 0.5  # 指标无差异时权重均分

    # 计算熵值与权重
    p = normalized_data / normalized_data.sum(axis=0, keepdims=True)
    p = np.clip(p, 1e-10, 1)  # 避免log(0)
    entropy = -np.sum(p * np.log(p), axis=0) / np.log(n)
    weights = (1 - entropy) / np.sum(1 - entropy)
    return weights.round(4).tolist()


def fuzzy_evaluation():
    # 1. 控制台输入配置
    problem = input("请输入当前求解的问题（如：产品第一层模糊评价）：")
    print(f"\n当前求解问题：{problem}")

    # 选择权重计算方式
    weight_choice = input("请选择权重计算方式（1=用Sheet1评价数据计算，2=用Sheet2单独数据计算）：")
    while weight_choice not in ["1", "2"]:
        weight_choice = input("输入错误，请重新选择（1或2）：")

    # 2. 读取Excel数据
    try:
        input_file = "fuzzy_input.xlsx"
        # 读取Sheet1：评价数据（必选）
        eval_df = pd.read_excel(input_file, sheet_name=0)
        objects = eval_df["评价对象"].tolist()
        indicators = eval_df.columns[1:].tolist()  # 指标名称（B列及以后）
        eval_data = eval_df[indicators].values  # 评价数据矩阵（对象×指标）

        # 根据选择读取权重数据源
        if weight_choice == "1":
            # 用Sheet1自身数据计算权重（评价对象作为样本）
            weight_data = eval_data  # 直接使用评价数据矩阵
            weight_source = "Sheet1的评价数据（评价对象作为样本）"
        else:
            # 用Sheet2数据计算权重（单独数据源）
            weight_data_df = pd.read_excel(input_file, sheet_name=1)
            weight_data = weight_data_df[indicators].values  # 确保指标名与Sheet1一致
            weight_source = "Sheet2的单独数据源"

        # 计算权重
        weights = entropy_weight(weight_data)
        print(f"权重计算来源：{weight_source}")
        print(f"自动计算的指标权重：{dict(zip(indicators, weights))}")
    except Exception as e:
        print(f"数据读取/权重计算失败：{str(e)}")
        return

    # 3. 梯形隶属函数参数（可自定义等级）
    mf_params = {
        "优": [80, 85, 100, 100],
        "良": [65, 70, 80, 85],
        "中": [50, 55, 65, 70],
        "差": [0, 0, 50, 55]
    }
    grades = list(mf_params.keys())

    # 4. 模糊评价计算
    result_list = []
    for i, obj in enumerate(objects):
        obj_data = eval_data[i]
        membership_matrix = [
            [trapezoidal_mf(data, *mf_params[grade]) for grade in grades]
            for data in obj_data
        ]

        # 加权合成
        final_membership = np.dot(weights, membership_matrix).round(4)
        final_grade = grades[np.argmax(final_membership)]

        # 整理结果
        result = {"评价对象": obj}
        for j, ind in enumerate(indicators):
            for k, grade in enumerate(grades):
                result[f"{ind}-{grade}"] = round(membership_matrix[j][k], 4)
        for k, grade in enumerate(grades):
            result[f"最终隶属度-{grade}"] = final_membership[k]
        result["最终评价等级"] = final_grade
        result_list.append(result)

    # 5. 输出Excel
    result_df = pd.DataFrame(result_list)
    now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    output_filename = f"{problem.replace('：', '-').replace(' ', '')}_{now}.xlsx"

    with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="评价结果", index=False)
        # 权重说明sheet
        ws_weight = writer.book.create_sheet("权重信息")
        weight_info = [
            ["权重计算方法", "熵权法"],
            ["权重数据来源", weight_source],
            ["样本数量", len(weight_data)],
            ["指标名称", "权重"],
            *[[ind, w] for ind, w in zip(indicators, weights)]
        ]
        for row_idx, (key, value) in enumerate(weight_info, 1):
            ws_weight.cell(row=row_idx, column=1, value=key)
            ws_weight.cell(row=row_idx, column=2, value=value)

    print(f"\n结果已保存至：{output_filename}")
    print("多层迭代提示：下一层计算时，将本次结果的「最终隶属度-各等级」作为新Sheet1的指标数据即可。")


if __name__ == "__main__":
    fuzzy_evaluation()
